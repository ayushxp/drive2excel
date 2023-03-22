"""import openpyxl
from openpyxl import load_workbook
data_file = 'C:/Users/Inno/Desktop/Links/input_01.xlsx'
# Load the entire workbook.
wb = load_workbook(data_file)
# Load one worksheet.
ws = wb['Sheet1']
all_rows = list(ws.rows)
sku = []
# Pull information from specific cells.
for row in all_rows[1:52]:
    state = row[0].value
    sku.append(state)
print(sku)"""


import openpyxl

"""from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

# Add a hyperlink


sheet.cell(row=2, column=1).hyperlink = "https://drive.google.com/file/d/1jCrP5L3Lyy8rA7qh5fSRMlY6Hm5l9VV6/view?usp=sharing"

sheet.cell(row=2, column=1).style = "Hyperlink"

wb.save("C:/Users/Inno/Desktop/Links/output_01.xlsx")

url=["https://drive.google.com/file/d/11f1NBteS_46dTPZHL6VfmfyXkfkOVAYm/view?usp=share_link"]
for i in list(url):
    x = i.split("/")
    l=x[-2]
    print(l)"""
#Url fetcher
"""from google.oauth2 import service_account
from getfilelistpy import getfilelist

SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'client_secret.json'
credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

resource = {
    "service_account": credentials,
    "id": "1VBpQyebKhSUDJKhKMeZX0zQz9fUNICgT",
    "fields": "files(name,id)",
}
res = getfilelist.GetFileList(resource)  # or r = getfilelist.GetFolderTree(resource)
print(res)"""
import os
from google.oauth2 import service_account
from getfilelistpy import getfilelist

SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'client_secret.json'
credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

resource = {
    "service_account": credentials,
    "id": "1YPfzAI3NVF30RmnAmL1c45mbBogfYp6k",
    "fields": "files(name,id)",
}
res = getfilelist.GetFileList(resource)
print(res)
#print(type(res))
all_keys = list(res.values())
all2=list(all_keys[1].values())
kost=list(all2[1])
print(kost)
dost=[]
for i in kost[1:]:
    dost.append(i)
print(dost)
import pandas as pd
# for i in dost:
#     pd.DataFrame(dost).to_csv("C:/Users/Inno/Desktop/output.csv", index=False,header=False)
df = pd.DataFrame(dost)
ui=list(all_keys[2])
print(ui)
kist=[]
k=1
for i in ui[1:]:
    i = ui[k]
    x1 = list(i.values())[0]
    kist.append(x1)
    k+=1

print(kist)

pot=[]
for i in kist:
    res1 = [sub['id'] for sub in i]
    #print(res)
    pot.append(res1)
print(pot)
for i in range(len(pot)):
    for j in range(len(pot[i])):
        str = pot[i][j]
        new_str = "https://drive.google.com/file/d/{}/view?usp=sharing".format(str)
        pot[i][j] = new_str
print(pot)
df2 = pd.DataFrame(pot)
result = pd.concat([df,df2], axis = 1)
#print(result)
pd.DataFrame(result).to_csv("C:/Users/Inno/Desktop/output.csv",index=False,mode='a',header=False)
os.system("C:/Users/Inno/Desktop/output.csv")


'''thisdict=ui[1]
thisdict1=ui[2]
thisdict2=ui[8]

x1 = list(thisdict.values())[0]
x2 = list(thisdict1.values())
x3 = list(thisdict2.values())
print(x1)
print(x2)
print(x3)'''

'''list1=[]
y=0
for po in x1:

    po=x1[y]

    first_val = list(po.values())[0]
    list1.append(f"https://drive.google.com/file/d/{first_val}/view?usp=sharing")
    y+=1
print(list1)'''

"""from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

# Add a hyperlink
t=1
for i in list1:
    sheet.cell(row=2, column=t).hyperlink=i
    sheet.cell(row=2, column=t).style = "Hyperlink"
    t+=1

wb.save("C:/Users/Inno/Desktop/Links/output_01.xlsx")"""






